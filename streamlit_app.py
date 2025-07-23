import pandas as pd
import re
import streamlit as st
import os
import openpyxl
import io
from openpyxl.styles import PatternFill

def extract_data_from_report(text_file):
    coverage_data = {}
    notest_data = {}
    pmsg_not_used = []
    pass_tests = {}  # Nouveau dictionnaire pour stocker les tests PASS
    component_test_counts = {}  # Pour compter le nombre de tests par composant. Ici on ne va compter que les composants ne passant qu'un seul test.
    
    text_content = text_file.getvalue().decode('utf-8', errors='ignore')
    
    # 1. Extraire les données de couverture (Test Summary)
    component_blocks = re.findall(r'Test Summary for ([A-Z]+\d+).*?Totals:.*?(\d+\.\d+)%', 
                                  text_content, re.DOTALL)
    
    for component, coverage in component_blocks:
        coverage_data[component] = float(coverage)
    
    # 2. Extraire les composants NOTEST et PMSG not used depuis la section Untested Devices
    untested_section = re.search(r'Untested Devices(.*?)(?=General Summary Report|\Z)', text_content, re.DOTALL)
    
    if untested_section:
        untested_content = untested_section.group(1)
        
        # Chercher les lignes avec "COMPONENT IS TESTED IN PARALLEL WITH" suivi de NOTEST
        notest_matches = re.findall(r'([A-Z]+\d+)\s+\(COMPONENT IS TESTED IN PARALLEL WITH ([A-Z]+\d+)\)\s+NOTEST', untested_content)
        for component, tested_with in notest_matches:
            comment = f"COMPONENT IS TESTED IN PARALLEL WITH {tested_with}"
            notest_data[component] = comment
        
        # Chercher les composants avec "PMSG is not used"
        pmsg_matches = re.findall(r'([A-Z]+\d+)\s+\(PMSG is not used\)', untested_content)
        pmsg_not_used.extend(pmsg_matches)
    
    # 3. Extraire les blocs de test avec PASS/FAIL
    test_blocks = re.findall(r'\*([A-Z]+\d+(?:_[A-Z]+\d*)*)\s+Units.*?(?:PASS|FAIL)\s', text_content, re.DOTALL)
    
    # Compter les occurrences de chaque composant principal
    for test_id in test_blocks:
        # Extrayez le composant principal (avant le premier underscore)
        main_component = test_id.split('_')[0]
        if main_component not in component_test_counts:
            component_test_counts[main_component] = 0
        component_test_counts[main_component] += 1
    
    # Trouver les composants qui n'ont qu'un seul test et qui sont PASS
    for test_block in re.finditer(r'\*([A-Z]+\d+(?:_[A-Z]+\d*)*)\s+Units.*?((?:PASS|FAIL))', text_content, re.DOTALL):
        test_id = test_block.group(1)
        result = test_block.group(2).strip()
        
        main_component = test_id.split('_')[0]
        if component_test_counts[main_component] == 1 and result.startswith("PASS"):
            pass_tests[main_component] = True
    
    return {
        "coverage": coverage_data,
        "notest": notest_data,
        "pmsg_not_used": pmsg_not_used,
        "pass_tests": pass_tests  # Ajout des tests PASS uniques
    }

def update_excel_with_data(excel_file, report_data):
    # Vérifier si le fichier est un CSV
    if excel_file.name.endswith('.csv'):
        # Lire le contenu du fichier pour vérifier s'il a un en-tête
        excel_file.seek(0)
        first_line = excel_file.readline().decode('utf-8', errors='ignore')
        excel_file.seek(0)
        
        # Si la première ligne ressemble à la nomenclature sans en-tête (format spécial)
        if first_line and not 'COMP.' in first_line and ',' in first_line:
            # Définir les noms de colonnes pour le format spécial
            column_names = ['COMP.', 'TYPE', 'VAL', 'TOL', 'STYLE', 'P/N', 'DESCRIPTION', 'LETTRE', 'CHIFFRE']
            df = pd.read_csv(excel_file, header=None, names=column_names)
            
            # Ajouter les colonnes manquantes
            if "BIBLIO" not in df.columns:
                df["BIBLIO"] = None
            if "STRATEGIE" not in df.columns:
                df["STRATEGIE"] = None
            if "STRUCTURAL" not in df.columns:
                df["STRUCTURAL"] = None
            if "COVERAGE %" not in df.columns:
                df["COVERAGE %"] = None
            if "PPVS" not in df.columns:
                df["PPVS"] = None
            if "REMARKS" not in df.columns:
                df["REMARKS"] = None
                
            # Réorganiser les colonnes pour correspondre au format demandé
            ordered_columns = ['COMP.', 'TYPE', 'STYLE', 'VAL', 'TOL', 'BIBLIO', 'P/N', 'DESCRIPTION',
                              'STRATEGIE', 'STRUCTURAL', 'PPVS', 'COVERAGE %', 'REMARKS']
            
            # S'assurer que toutes les colonnes existent
            for col in ordered_columns:
                if col not in df.columns:
                    df[col] = None
            
            # Réordonner les colonnes et supprimer LETTRE et CHIFFRE
            df = df[ordered_columns]
        else:
            # CSV standard avec en-tête
            df = pd.read_csv(excel_file)
    else:
        # Fichier Excel standard
        df = pd.read_excel(excel_file)
    
    # S'assurer que les colonnes nécessaires existent
    if "COVERAGE %" not in df.columns:
        df["COVERAGE %"] = None
    if "PPVS" not in df.columns:
        df["PPVS"] = None
    if "REMARKS" not in df.columns:
        df["REMARKS"] = None
    
    coverage_data = report_data["coverage"]
    notest_data = report_data["notest"]
    pmsg_not_used = report_data["pmsg_not_used"]
    pass_tests = report_data.get("pass_tests", {})  # Récupérer les tests PASS uniques
    
    # Créer un dictionnaire pour stocker les informations de formatage
    format_info = {}
    
    # Créer des ensembles pour suivre les composants déjà traités dans le rapport
    processed_components = set()
    
    for index, row in df.iterrows():
        if "COMP." not in row or pd.isna(row["COMP."]):
            continue
            
        component = row["COMP."]
        
        # 1. Traitement des composants avec couverture
        if component in coverage_data:
            df.at[index, "COVERAGE %"] = f"{coverage_data[component]:.2f}%"
            df.at[index, "PPVS"] = "OK"
            # Mémoriser les cellules à formater en vert
            format_info[index] = {"column": "PPVS", "color": "green"}
            processed_components.add(component)
            
        # 2. Traitement des composants NOTEST (SOUS-TEST) - composants testés en parallèle
        elif component in notest_data:
            df.at[index, "PPVS"] = "SOUS-TEST"
            df.at[index, "REMARKS"] = notest_data[component]
            # Laisser la colonne COVERAGE % inchangée si elle a déjà une valeur
            if pd.isna(df.at[index, "COVERAGE %"]) or df.at[index, "COVERAGE %"] is None:
                df.at[index, "COVERAGE %"] = None
            # Mémoriser les cellules à formater en jaune clair
            format_info[index] = {"column": "PPVS", "color": "yellow"}
            processed_components.add(component)
            
        # 3. Traitement des composants PMSG not used (NOTEST)
        elif component in pmsg_not_used:
            df.at[index, "PPVS"] = "NOTEST"
            # Laisser la colonne COVERAGE % inchangée si elle a déjà une valeur
            if pd.isna(df.at[index, "COVERAGE %"]) or df.at[index, "COVERAGE %"] is None:
                df.at[index, "COVERAGE %"] = None
            # Mémoriser les cellules à formater en rouge clair
            format_info[index] = {"column": "PPVS", "color": "red"}
            processed_components.add(component)
        
        # 4. Traitement des composants avec un test unique PASS
        elif component in pass_tests:
            df.at[index, "PPVS"] = "OK"
            # Mémoriser les cellules à formater en vert
            format_info[index] = {"column": "PPVS", "color": "green"}
            processed_components.add(component)
    
    # Afficher les statistiques des données traitées
    st.write(f"Composants avec coverage trouvés: {len(coverage_data)}")
    st.write(f"Composants SOUS-TEST trouvés: {len(notest_data)}")
    st.write(f"Composants NOTEST (PMSG not used) trouvés: {len(pmsg_not_used)}")
    st.write(f"Composants avec test unique PASS trouvés: {len(pass_tests)}")
    st.write(f"Composants traités dans le tableau Excel: {len(processed_components)}")
    
    return df, format_info

def main():
    st.set_page_config(page_title="Coverage Excel Update", layout="wide")
    
    st.title("Coverage Excel Update")
    
    st.markdown("""
    Cette application permet de mettre à jour un fichier Excel avec des données de couverture de test.
    """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Sélectionner le fichier Excel/CSV")
        excel_file = st.file_uploader("Choisissez un fichier Excel ou CSV", type=["xlsx", "csv"])
        
    with col2:
        st.subheader("Sélectionner le rapport de couverture")
        text_file = st.file_uploader("Choisissez un fichier texte de rapport", type=None, accept_multiple_files=False)
    
    # Afficher des informations sur les fichiers acceptés
    st.info("✅ Formats acceptés pour le rapport: tous les types de fichiers texte (y compris sans extension)")
    
    if excel_file is not None and text_file is not None:
        import tempfile
        
        # Afficher le nom du fichier texte sélectionné
        file_name = text_file.name if hasattr(text_file, 'name') else "Fichier sans nom"
        st.write(f"📄 Fichier de rapport sélectionné: **{file_name}**")
        st.success("Fichier de rapport chargé avec succès, peu importe son extension.")
            
        if st.button("Traiter les fichiers", type="primary"):
            with st.spinner("Traitement des fichiers en cours..."):
                try:
                    # Réinitialiser le curseur du fichier texte
                    text_file.seek(0)
                    report_data = extract_data_from_report(text_file)
                    
                    # Afficher un résumé des données extraites
                    st.write("Données extraites du rapport:")
                    st.write(f"- Composants avec couverture: {len(report_data['coverage'])}")
                    st.write(f"- Composants NOTEST (SOUS-TEST): {len(report_data['notest'])}")
                    st.write(f"- Composants PMSG not used (NOTEST): {len(report_data['pmsg_not_used'])}")
                    st.write(f"- Composants avec test unique PASS: {len(report_data.get('pass_tests', {}))}")
                    
                    # Vérification plus stricte des données
                    if not report_data["coverage"] and not report_data["notest"] and not report_data["pmsg_not_used"] and not report_data.get("pass_tests", {}):
                        st.error("Aucune donnée trouvée dans le fichier texte. Assurez-vous que le format du fichier est correct.")
                        st.info("⚠️ Le fichier doit contenir des sections 'Test Summary for' ou 'Untested Devices' avec des données de couverture.")
                    else:
                        # Réinitialiser le curseur du fichier Excel
                        excel_file.seek(0)
                        updated_df, format_info = update_excel_with_data(excel_file, report_data)
                        
                        st.success(f"Traitement terminé!")
                        
                        st.subheader("Aperçu du fichier mis à jour")
                        st.dataframe(updated_df)
                        
                        import io
                        import openpyxl
                        from openpyxl.styles import PatternFill, Alignment, Font
                        
                        # Transformer le DataFrame en Excel
                        # En utilisant pandas pour écrire dans un BytesIO
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            # Écrire les données dans l'onglet "Nomenclature"
                            updated_df.to_excel(writer, sheet_name="Nomenclature", index=False)
                            
                            # Accéder au classeur et à la feuille de travail
                            workbook = writer.book
                            worksheet = writer.sheets["Nomenclature"]
                            
                            # Formater les en-têtes
                            header_font = Font(bold=True, color="FFFFFF", name="Aptos Narrow", italic=True)  # Blanc, Aptos Narrow, Italique
                            header_fill = PatternFill(start_color="00AA91", end_color="00AA91", fill_type="solid")  # Turquoise #00AA91
                            header_alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Style de remplissage pour chaque couleur
                            green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Vert clair pour OK
                            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Jaune pour SOUS-TEST
                            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge pour NOTEST
                            
                            # Styles des bordures pour le quadrillage
                            from openpyxl.styles import Border, Side
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            
                            # Initialiser les indices des colonnes importantes
                            coverage_col_idx = None
                            ppvs_col_idx = None
                            remarks_col_idx = None
                            
                            # Appliquer le formatage aux en-têtes et trouver les indices des colonnes importantes
                            for col_num, column_title in enumerate(updated_df.columns, 1):
                                cell = worksheet.cell(row=1, column=col_num)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                                cell.border = thin_border
                                
                                # Déterminer les indices des colonnes importantes
                                if column_title == "COVERAGE %":
                                    coverage_col_idx = col_num
                                elif column_title == "PPVS":
                                    ppvs_col_idx = col_num
                                elif column_title == "REMARKS":
                                    remarks_col_idx = col_num
                            
                            # Appliquer le formatage aux données
                            for row_idx in range(len(updated_df)):
                                for col_num in range(1, len(updated_df.columns) + 1):
                                    # Appliquer les bordures à toutes les cellules
                                    cell = worksheet.cell(row=row_idx+2, column=col_num)
                                    cell.border = thin_border
                                    cell.font = Font(name="Aptos Narrow")  # Appliquer la police Aptos Narrow à toutes les cellules
                                    
                                # Formater la colonne COVERAGE %
                                if coverage_col_idx is not None:
                                    coverage_cell = worksheet.cell(row=row_idx+2, column=coverage_col_idx)
                                    if coverage_cell.value and isinstance(coverage_cell.value, str) and "%" in str(coverage_cell.value):
                                        try:
                                            coverage_cell.value = float(str(coverage_cell.value).replace("%", "")) / 100
                                            coverage_cell.number_format = '0.00%'
                                            coverage_cell.alignment = Alignment(horizontal='center')
                                        except ValueError:
                                            pass
                                
                                # Appliquer l'alignement à la colonne PPVS (mais pas les couleurs, elles seront gérées par le formatage conditionnel)
                                if ppvs_col_idx is not None and row_idx in format_info:
                                    ppvs_cell = worksheet.cell(row=row_idx+2, column=ppvs_col_idx)
                                    ppvs_cell.alignment = Alignment(horizontal='center')
                                    # Nous ne définissons plus de couleurs ici, elles seront gérées par le formatage conditionnel
                            
                            # Ajuster la largeur des colonnes
                            from openpyxl.utils import get_column_letter
                            for i, column in enumerate(updated_df.columns):
                                # Trouver la longueur maximale dans la colonne
                                max_length = max(
                                    updated_df[column].astype(str).map(lambda x: len(str(x)) if not pd.isna(x) else 0).max(),
                                    len(str(column))
                                )
                                # Limiter entre 10 et 30 caractères
                                col_width = min(max(max_length + 2, 10), 30)
                                col_letter = get_column_letter(i+1)
                                worksheet.column_dimensions[col_letter].width = col_width
                                
                            # Créer l'onglet "Liste" pour les options des menus déroulants
                            worksheet_liste = workbook.create_sheet("Liste")
                            
                            # Définir les valeurs uniques pour chaque colonne
                            dropdown_columns = {
                                "TYPE": ["ANALOG", "CAP", "CONN", "DIODE", "FUSE", "HYBRID", "IND", "JUMPER", "LED", 
                                         "LOGIC", "NPN", "NFET", "NJFET", "OTHER", "PCAP", "PFET", "PJFET", "PNP", 
                                         "RES", "VRES", "ZENER"],
                                "STYLE": ["1N", "10N", "47N", "47P", "1U", "15U", "3.3U"],
                                "BIBLIO": ["Teradyne", "Adaptée", "NOUVEAU", "Sans"],
                                "STRATEGIE": ["Analog", "Analog PWR", "Hybride", "Logic", "Logic + PROG", "Cluster", 
                                             "Mesure I", "Mesure V", "Mesure F", "Fonctionnel"],
                                "STRUCTURAL": ["Junction", "Capacitive", "Both"],
                                "PPVS": ["OK", "SOUS-TEST", "NOTEST"]
                            }
                            
                            # Style pour les en-têtes de l'onglet Liste
                            liste_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                            liste_header_font = Font(bold=True, color="FFFFFF", name="Aptos Narrow", italic=True)  # Même style que l'onglet principal
                            
                            # Remplir l'onglet Liste
                            col_idx = 1
                            for col_name, values in dropdown_columns.items():
                                # En-tête
                                header_cell = worksheet_liste.cell(row=1, column=col_idx)
                                header_cell.value = col_name
                                header_cell.font = liste_header_font
                                header_cell.fill = liste_header_fill
                                header_cell.alignment = Alignment(horizontal='center')
                                header_cell.border = thin_border
                                
                                # Valeurs
                                for row_idx, value in enumerate(values, start=2):
                                    cell = worksheet_liste.cell(row=row_idx, column=col_idx)
                                    cell.value = value
                                    cell.border = thin_border
                                    cell.font = Font(name="Aptos Narrow")  # Appliquer la police Aptos Narrow
                                
                                # Ajuster largeur
                                max_width = max(len(col_name), max(len(str(v)) for v in values) if values else 0)
                                worksheet_liste.column_dimensions[get_column_letter(col_idx)].width = max_width + 4
                                
                                col_idx += 1
                                
                            # Ajouter des validations de données pour créer des menus déroulants
                            from openpyxl.worksheet.datavalidation import DataValidation
                            
                            # Appliquer les couleurs aux valeurs PPVS dans l'onglet Liste
                            if "PPVS" in dropdown_columns:
                                ppvs_col_idx = list(dropdown_columns.keys()).index("PPVS") + 1
                                ppvs_col_letter = get_column_letter(ppvs_col_idx)
                                
                                # Appliquer les couleurs aux valeurs dans l'onglet Liste
                                for row_idx, value in enumerate(dropdown_columns["PPVS"], start=2):
                                    cell = worksheet_liste.cell(row=row_idx, column=ppvs_col_idx)
                                    if value == "OK":
                                        cell.fill = green_fill
                                    elif value == "SOUS-TEST":
                                        cell.fill = yellow_fill
                                    elif value == "NOTEST":
                                        cell.fill = red_fill
                                    cell.alignment = Alignment(horizontal='center')
                                    cell.font = Font(name="Aptos Narrow")  # Appliquer la police Aptos Narrow
                            
                            # Pour chaque colonne avec menu déroulant
                            for i, col_name in enumerate(dropdown_columns.keys()):
                                if col_name in updated_df.columns:
                                    # Trouver l'indice de la colonne dans le DataFrame
                                    col_idx = list(updated_df.columns).index(col_name) + 1
                                    col_letter = get_column_letter(col_idx)
                                    
                                    # Créer la formule pour le menu déroulant
                                    list_col_letter = get_column_letter(i + 1)
                                    max_row = len(dropdown_columns[col_name]) + 1
                                    formula = f"Liste!${list_col_letter}$2:${list_col_letter}${max_row}"
                                    
                                    # Créer la validation
                                    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                                    dv.add(f"{col_letter}2:{col_letter}{len(updated_df)+1}")
                                    worksheet.add_data_validation(dv)
                                    
                                    # Si c'est la colonne PPVS, ajouter un formatage conditionnel pour changer les couleurs automatiquement
                                    if col_name == "PPVS":
                                        from openpyxl.formatting.rule import CellIsRule
                                        
                                        # Règle pour OK - couleur verte
                                        green_rule = CellIsRule(operator='equal', formula=['"OK"'], stopIfTrue=True, fill=green_fill)
                                        worksheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(updated_df)+1}", green_rule)
                                        
                                        # Règle pour SOUS-TEST - couleur jaune
                                        yellow_rule = CellIsRule(operator='equal', formula=['"SOUS-TEST"'], stopIfTrue=True, fill=yellow_fill)
                                        worksheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(updated_df)+1}", yellow_rule)
                                        
                                        # Règle pour NOTEST - couleur rouge
                                        red_rule = CellIsRule(operator='equal', formula=['"NOTEST"'], stopIfTrue=True, fill=red_fill)
                                        worksheet.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(updated_df)+1}", red_rule)
                        
                        # Terminer l'écriture du fichier Excel
                        output.seek(0)
                        
                        # Déterminer le nom du fichier de sortie
                        if excel_file.name.endswith('.csv'):
                            file_name = excel_file.name.replace('.csv', '_updated.xlsx')
                        else:
                            file_name = excel_file.name.replace('.xlsx', '_updated.xlsx')
                        
                        # Bouton de téléchargement pour le fichier formaté
                        st.download_button(
                            label="Télécharger le fichier Excel mis à jour",
                            data=output,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                except Exception as e:
                    st.error(f"Erreur lors du traitement: {str(e)}")
        
        st.text("Note: Les fichiers sont traités directement en mémoire et ne sont pas sauvegardés sur le serveur.")

if __name__ == "__main__":
    main()
